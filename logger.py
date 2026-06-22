import os
import hashlib
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = "inventory_log.xlsx"
COLUMNS = ["Timestamp", "Shelf", "Cube Count", "Cylinder Count", "Small Cups Count", "Total Count"]

# All 7 classes mapped to log column
CLASS_BASE = {
    "bottle":     "cylinder",
    "chips":      "cube",
    "cubes":      "cube",
    "cups":       "cylinder",
    "cylinders":  "cylinder",
    "small_cups": "small_cups",
    "tin can":    "cylinder",
}

# ── Stability filter ───────────────────────────────────────────────────────
# A shelf is only logged after its counts stay identical for this many
# consecutive calls. At INFER_EVERY=1s, STABLE_FRAMES=3 means 3 seconds
# of identical readings before writing. Kills the flicker problem.
STABLE_FRAMES = 3

_log         = []
_last_hashes = {}   # shelf → last hash that was actually written to Excel
_pending     = {}   # shelf → {"hash": str, "count": int, "row": dict}


def _shelf_hash(shelf_id, cube, cylinder, small_cups):
    payload = f"{shelf_id}:{cube}:{cylinder}:{small_cups}"
    return hashlib.md5(payload.encode()).hexdigest()


def _count_classes(items: dict) -> tuple:
    cubes = cylinders = small_cups = 0
    for cls, count in items.items():
        base = CLASS_BASE.get(cls)
        if base == "cube":
            cubes += count
        elif base == "cylinder":
            cylinders += count
        elif base == "small_cups":
            small_cups += count
        else:
            print(f"[LOGGER] Warning: unknown class '{cls}' — skipped")
    return cubes, cylinders, small_cups


def _init_excel():
    if os.path.exists(EXCEL_PATH):
        return
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Log"
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center")
    for i, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)


def _append_rows(rows: list[dict]):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in rows:
        ws.append([
            row["Timestamp"],
            row["Shelf"],
            row["Cube Count"],
            row["Cylinder Count"],
            row["Small Cups Count"],
            row["Total Count"],
        ])
    wb.save(EXCEL_PATH)


def log_inventory(inventory_dict: dict):
    """
    Called every inference cycle with the full inventory.

    Rules:
      1. Skip shelves where total == 0 AND last logged state was also 0
         (no point recording "still empty").
      2. Only write to Excel after a shelf's counts have been IDENTICAL
         for STABLE_FRAMES consecutive calls — eliminates YOLO flicker.
    """
    _init_excel()

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows_to_write = []

    for shelf_id, items in inventory_dict.items():
        cubes, cylinders, small_cups = _count_classes(items)
        total = cubes + cylinders + small_cups

        # Rule 1 — skip zero→zero transitions silently
        last_written = _last_hashes.get(shelf_id)
        if total == 0 and last_written is None:
            continue   # never been non-zero, nothing to record

        current_hash = _shelf_hash(shelf_id, cubes, cylinders, small_cups)

        # Already written this exact state — nothing changed
        if current_hash == last_written:
            _pending.pop(shelf_id, None)
            continue

        # Build candidate row
        entry = {
            "Timestamp":        timestamp,
            "Shelf":            shelf_id,
            "Cube Count":       cubes,
            "Cylinder Count":   cylinders,
            "Small Cups Count": small_cups,
            "Total Count":      total,
        }

        # Rule 2 — stability filter
        p = _pending.get(shelf_id)
        if p and p["hash"] == current_hash:
            p["count"] += 1
        else:
            # New candidate — start counter
            _pending[shelf_id] = {"hash": current_hash, "count": 1, "row": entry}
            continue

        if p["count"] >= STABLE_FRAMES:
            # Stable — commit to Excel
            _last_hashes[shelf_id] = current_hash
            _pending.pop(shelf_id, None)
            _log.append(entry)
            rows_to_write.append(entry)
            print(f"[LOG] {timestamp}  Shelf {shelf_id}  "
                  f"cubes={cubes}  cylinders={cylinders}  "
                  f"small_cups={small_cups}  total={total}")

    if rows_to_write:
        _append_rows(rows_to_write)


def get_log() -> list[dict]:
    return list(_log)


def get_log_df() -> pd.DataFrame:
    return pd.DataFrame(_log, columns=COLUMNS)


def clear_log():
    _log.clear()
    _last_hashes.clear()
    _pending.clear()